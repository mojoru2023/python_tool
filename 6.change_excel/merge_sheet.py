import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# 定义文件路径
a_file = "a.xlsx"  # 修改为实际的 a.xlsx 文件路径
b_file = "b.xlsx"  # 修改为实际的 b.xlsx 文件路径
mojoru_file = "mojoru.xlsx"  # 修改为输出的 mojoru.xlsx 文件路径


# 复制单元格样式的函数
def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = source_cell.font.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.alignment = source_cell.alignment.copy()
        target_cell.protection = source_cell.protection.copy()
        target_cell.number_format = source_cell.number_format


# 1. 复制 a.xlsx 为 mojoru.xlsx，并修改 Sheet2 中 aaa 后面的内容为 mojoru
wb_a = load_workbook(a_file)
wb_mojoru = Workbook()

# 复制第一张表（默认创建的 Sheet）到 mojoru.xlsx
sheet1_a = wb_a.active
sheet1_mojoru = wb_mojoru.active
sheet1_mojoru.title = sheet1_a.title

for row in sheet1_a.iter_rows(values_only=False):
    for cell in row:
        new_cell = sheet1_mojoru[cell.coordinate]
        new_cell.value = cell.value
        copy_cell_style(cell, new_cell)

# 处理 Sheet2  修改时有问题
if 'Sheet2' in wb_a.sheetnames:
    sheet2_a = wb_a['Sheet2']
    sheet2_mojoru = wb_mojoru.create_sheet(title='Sheet2')

    for row in range(1, sheet2_a.max_row + 1):
        for col in range(1, sheet2_a.max_column + 1):
            cell_value = sheet2_a.cell(row=row, column=col).value
            new_cell = sheet2_mojoru.cell(row=row, column=col)

            # 复制当前单元格的值和样式
            new_cell.value = cell_value
            copy_cell_style(sheet2_a.cell(row=row, column=col), new_cell)

            # 如果当前单元格为 "aaa"，则将该行后续单元格的值修改为 "mojoru"
            if cell_value == "aaa":
                sheet2_mojoru.cell(row=row, column=col+ 1).value = "mojoru"



# 保存 mojoru.xlsx
wb_mojoru.save(mojoru_file)

# 2. 读取 b.xlsx 中的 Sheet3 的内容、样式和图片，复制到 mojoru.xlsx 中
wb_b = load_workbook(b_file)

if 'Sheet3' in wb_b.sheetnames:
    sheet3_b = wb_b['Sheet3']
    sheet3_mojoru = wb_mojoru.create_sheet(title='mojoru-sheet')

    # 复制单元格内容和样式
    for row in sheet3_b.iter_rows():
        for cell in row:
            new_cell = sheet3_mojoru.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                copy_cell_style(cell, new_cell)

    # 复制图片
    for img in sheet3_b._images:
        new_image = Image(img.ref)
        new_image.anchor = img.anchor
        sheet3_mojoru.add_image(new_image)

# 3. 按照指定顺序重新排列 Sheet
sheets = wb_mojoru.sheetnames
if 'mojoru-sheet' in sheets:
    wb_mojoru.move_sheet('mojoru-sheet', offset=-len(sheets) + 1)

# 保存最终的 mojoru.xlsx
wb_mojoru.save(mojoru_file)

print("处理完成，mojoru.xlsx 已生成。")
