import os
import shutil
from openpyxl import load_workbook
from datetime import datetime

# 定义文件夹路径
input_folder = "C:\\Users\\user\\Desktop\\service-"  # 修改为实际的输入文件夹路径
aaa_folder = os.path.join(input_folder, 'aaa')
bbb_folder = os.path.join(input_folder, 'bbb')
output_folder = "C:\\Users\\user\\Desktop\\service-\\output"  # 修改为实际的输出文件夹路径

# 创建output文件夹（如果不存在）
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# 读取aaa和bbb文件夹里面的所有文件名
aaa_files = [f for f in os.listdir(aaa_folder) if f.endswith('.xlsx')]
bbb_files = [f for f in os.listdir(bbb_folder) if f.endswith('.xlsx')]

# 将aaa文件夹里面的所有excel复制一份到output文件夹中
for file in aaa_files:
    src = os.path.join(aaa_folder, file)
    dst = os.path.join(output_folder, file)
    shutil.copy(src, dst)

    # 修改excel文件
    workbook = load_workbook(dst)

    # 读取所有sheet名的信息
    sheet_names = workbook.sheetnames
    print(f"Processing {file}: Sheets - {sheet_names}")

    # 针对sheet2，修改A列和N列右边的单元格内容为mojoru
    if len(sheet_names) >= 2:
        sheet2 = workbook[sheet_names[1]]  # sheet2

        for row in range(1, sheet2.max_row + 1):  # 遍历行
            # 检查A列（即第一列）
            if sheet2.cell(row=row, column=1).value == "aaa":  # A列
                right_cell = sheet2.cell(row=row, column=2)  # 右侧单元格
                right_cell.value = "mojoru"

            # 检查N列（以N为列索引的列）
            N = 14  # 假设N列为第14列（即N列）
            if sheet2.cell(row=row, column=N).value == "aaa":
                right_cell_N = sheet2.cell(row=row, column=N + 1)  # N列右侧单元格
                right_cell_N.value = "mojoru"

    # 将sheet3的sheet名修改为mojoru-test
    if len(sheet_names) >= 3:
        sheet3 = workbook[sheet_names[2]]  # sheet3
        sheet3.title = "mojoru-test"

    # 保存更改
    workbook.save(dst)

# 修改output里面的excel的文件名，增加当前天的日期
today_date = datetime.now().strftime('%Y-%m-%d')
for file in os.listdir(output_folder):
    if file.endswith('.xlsx'):
        new_file_name = file.replace('.xlsx', f'_{today_date}.xlsx')
        os.rename(os.path.join(output_folder, file), os.path.join(output_folder, new_file_name))
